import openpyxl as xl
from openpyxl.worksheet.worksheet import Worksheet
from typing import List, Union


def highlighter(worksheet: Worksheet, color_hex: str, *cells: Union[str, List[str]]) -> None:
    """ Highlights excel cell(s) using openpyxl

    :param worksheet: (var: openpyxl.worksheet.worksheet) Your worksheet, as designated by openpyxl library
    :param cells: (str) The cell(s) you wish to highlight
    :param color_hex: (str) The 6 digit hex code of the color you wish to use to highlight

    Note:
        Colors such as "blue" are not accepted. Please use something like '088F8F' instead

    Example;
        workbook = load_workbook(file_path)
        worksheet = workbook['excel_file']

        highlighter(worksheet,'088F8F', *['A2', 'B2', 'C2'])
        highlighter(worksheet,'088F8F', D2)

    """

    marker = xl.styles.PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")

    for cell in cells:
        worksheet[cell].fill = marker


def bold(worksheet: Worksheet, *cells: Union[str, List[str]]) -> None:
    """ Formats excel cell(s) with bold font.

    :param worksheet: (var: openpyxl.worksheet.worksheet) Your worksheet, as designated by openpyxl library
    :param cells: (str) The cell(s) you wish to make bold

    Example;
        workbook = load_workbook(file_path)
        worksheet = workbook['excel_file']

        bold(worksheet,['A2', 'B2', 'C2])
    """

    bold_bigger_font = xl.styles.Font(bold=True)

    for cell in cells:
        worksheet[cell].font = bold_bigger_font


def currency_formatter(worksheet: Worksheet, *cells: Union[str, List[str]]) -> None:
    """ Formats numerical excel cell(s) as a currency.

    :param worksheet: (var: openpyxl.worksheet.worksheet) Your worksheet, as designated by openpyxl library
    :param cells: The cell(s) to which you wish to add currency formatting

    Example:
        wb = xl.load_workbook(excel_file_path)
        ws = wb[sheet_name]

        currency_formatter(ws, ['C2','B2'])
    """
    for cell in cells:
        worksheet[cell].number_format = '$#,##0.00'


def thousands_sep_formatter(worksheet: Worksheet, *cells: Union[str, List[str]]) -> None:
    """ Formats numerical excel cell(s) with thousands separator.

    :param worksheet: (var: openpyxl.worksheet.worksheet) Your worksheet, as designated by openpyxl library
    :param cells: The cell(s) to which you wish to add currency formatting

    Example:
        wb = xl.load_workbook(excel_file_path)
        ws = wb[sheet_name]

        thousands_sep_formatter(ws, 'C2')
    """

    for cell in cells:
        worksheet[cell].number_format = '#,##0'

def percent_formatter(worksheet: Worksheet, *cells: Union[str, List[str]]) -> None:
    """ Formats numerical excel cell(s) as a percentage, rounded to the nearest whole number.

    :param worksheet: (var: openpyxl.worksheet.worksheet) Your worksheet, as designated by openpyxl library
    :param cells: The cell(s) to which you wish to apply formatting to

    Example:
        wb = xl.load_workbook(excel_file_path)
        ws = wb[sheet_name]

        percent_formatter(ws, 'C2')
    """

    for cell in cells:
        worksheet[cell].number_format = '#%'


def remove_gridlines(worksheet: Worksheet) -> None:
    """Removes gridlines from an excel worksheet

    :param worksheet: (var: openpyxl.worksheet.worksheet)
                      Specifies the openpyxl worksheet from which you wish to remove gridlines

    Example:
        wb = xl.load_workbook(excel_file_path)
        ws = wb[sheet_name]

        remove_gridlines(ws)
    """
    worksheet.sheet_view.showGridLines = False



def freeze_first_row(worksheet: Worksheet, row_num: int) -> None:
    """
    Freezes the headers row in an excel worksheet.

    :param worksheet (var:openpyxl): the worksheet of an openpyxl workbook

    Example:
        wb = xl.load_workbook(excel_file_path)
        ws = wb[sheet_name]

        freeze_first_col(ws)
    """
    worksheet.freeze_panes = f'A{row_num}'
