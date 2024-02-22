import openpyxl as xl
from openpyxl.worksheet.worksheet import Worksheet


def excel_table_creator(worksheet: Worksheet, table_name: str) -> None:
    """ Creates an excel table out of an array using the openpyxl library.

    :param worksheet: (var: Worksheet) your openpyxl worksheet
    :param table_name: (str) assign your table a name
    :return: your excel array will become a basic default blue table.

    Notes:
        -Be sure to import your openpyxl workbook and define worksheets beforehand.
        -The array range is assumed by taking the last row of the last column. The starting value is always A1.
        -Excel doesn't allow duplicate table names in a worksheet, so be sure to assign unique str values to table_name.

    Example:
        wb = xl.load_workbook(excel_file_path)
        ws = wb[sheet_name]

        excel_table_creator(ws, 'Table1')
    """
    if not isinstance(worksheet, Worksheet):
        raise ValueError("The 'worksheet' parameter must be an openpyxl worksheet. Please define a wb and ws first.")

    last_column = xl.utils.get_column_letter(len(list(worksheet.columns)))
    last_row = worksheet.max_row

    start_of_table = "A1"
    end_of_table = f"{last_column}{last_row}"
    table_range = f"{start_of_table}:{end_of_table}"

    table_design = xl.worksheet.table.TableStyleInfo(
        name="TableStyleMedium9"
        , showFirstColumn=False
        , showLastColumn=False
        , showRowStripes=True
        , showColumnStripes=False
    )

    create_table = xl.worksheet.table.Table(
        displayName=f"Table{table_name}"
        , ref=table_range
    )

    create_table.tableStyleInfo = table_design

    worksheet.add_table(create_table)
